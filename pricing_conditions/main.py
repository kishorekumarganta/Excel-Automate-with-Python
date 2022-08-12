import pandas as pd
import openpyxl
from os import listdir
from os.path import isfile, join
import datetime

# import xlrd
# from openpyxl.workbook import Workbook
# from openpyxl.reader.excel import load_workbook, InvalidFileException

xlsxFolderPath = "xlsx"
xlsxFolderPathRaw = "raw_xlsx"

# xlsxFolderPath = "test\\xlsx"
# xlsxFolderPathRaw = "test\\raw_xlsx"

def getFilesFromDirectory(selectedPath):
    return [f for f in listdir(selectedPath) if isfile(join(selectedPath, f))]

filesToHandle = getFilesFromDirectory(xlsxFolderPathRaw)

def getConditionValue(sheet):
    condition = sheet.cell(1, 1).value
    return condition.split(": ")[1].strip()

mainDF = pd.DataFrame() # Create empty dataframe

# def open_xls_as_xlsx(filename):
#     # first open using xlrd
#     book = xlrd.open_workbook(filename)
#     index = 0
#     nrows, ncols = 0, 0
#     while nrows * ncols == 0:
#         sheet = book.sheet_by_index(index)
#         nrows = sheet.nrows
#         ncols = sheet.ncols
#         index += 1

#     # prepare a xlsx sheet
#     book1 = Workbook()
#     sheet1 = book1.get_active_sheet()

#     for row in xrange(0, nrows):
#         for col in xrange(0, ncols):
#             sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

#     return book1

for filename in filesToHandle:

    # Get condition value and delete first row
    print("Start processing {}...".format(filename))
    wb = openpyxl.load_workbook(xlsxFolderPathRaw + "\\" + filename)
    # wb = open_xls_as_xlsx(xlsxFolderPathRaw + "\\" + filename)
    sheet = wb.worksheets[0]
    conditionValue = getConditionValue(sheet)
    print("Condition value {}".format(conditionValue))
    print("Start removing first row and saving file...")
    sheet.delete_rows(1, 1)
    newFilePath = xlsxFolderPath + "\\" + filename
    wb.save(newFilePath)
    print("First row deleted and file saved at {}".format(newFilePath))

    # Get dataframe from first sheet
    print("Start reading file at {}...".format(newFilePath))
    xlsx = pd.ExcelFile(newFilePath)
    df = pd.read_excel(xlsx, sheet_name=0)
    print(df.dtypes)

    # Normalize column names
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_').str.replace('(', '').str.replace(')', '')
    colNames = df.columns.tolist()
    print("Column names have been normalized {}".format(colNames))

    # Add column with the conditionValue
    dfSize = len(df)
    newCol = [conditionValue for x in range(dfSize)]
    colName = "condition_value"
    df[colName] = newCol
    print("Condition value has been added to column {}".format(colName))

    # Add column with the filename
    dfSize = len(df)
    newCol = [filename for x in range(dfSize)]
    colName = "filename"
    df[colName] = newCol
    print("Filename value has been added to column {}".format(colName))

    # Standardize names
    if "inv._pty" in colNames:
        df['vendor_id'] = df['inv._pty'].astype(str).str.strip() # Remove unecessary spaces
    else:
        df['vendor_id'] = df['vendor'].astype(str).str.strip() # Remove unecessary spaces
    if "material" in colNames:
        df['material_id'] = df['material']
    else:
        df['material_id'] = df['matl_group']
    
    # Create product code with concatenation
    def getProductCode(row):
        return str(row['vendor_id']) + str(row['material_id'])
    colName = "product_code"
    df[colName] = df.apply (lambda row: getProductCode(row), axis=1)
    print("Product code has been added to column {}".format(colName))

    # Create new cost per unit
    def getCostPerUnit(row):
        if (not ((isinstance(row['amount'], float) or isinstance(row['amount'], int)) and (isinstance(row['per'], float) or isinstance(row['per'], int)))):
            return 0
        return float(row['amount']) / float(row['per'])
    colName = "cost_per_unit"
    df[colName] = df.apply (lambda row: getCostPerUnit(row), axis=1)
    print("Cost per unit has been added to column {}".format(colName))

    mainDF = mainDF.append(df)

print("Main dataframe is now complete")
print(mainDF)

# Convert the content of the valid_from column to a date type
print("Converting valid_from column to type date")
mainDF['valid_from_dt'] = pd.to_datetime(mainDF['valid_from'], errors='coerce') # doc on pd.to_datetime https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.to_datetime.html
print(mainDF.dtypes)
print("Dataframe with invalid dates")
invalidDatesDF = mainDF[mainDF['valid_from_dt'].isnull()]
print(invalidDatesDF)

# Split dataframe based on valid_from future dates
dateToLimit = datetime.datetime(2021,12,31)
# dateToLimit = datetime.datetime.now() # now
futureDatesDF = mainDF[(mainDF['valid_from_dt'] > dateToLimit)]
keepDF = mainDF[(mainDF['valid_from_dt'] < dateToLimit)]
print("Dataframe with future dates")
print(futureDatesDF)

# Remove duplicates
keepDF = keepDF.dropna(subset=['valid_from_dt']) # Remove rows where the conversion did not work
keepDF = keepDF.sort_values(by=['valid_from_dt'], ascending=False)
keepDF = keepDF.drop_duplicates(subset ="product_code", keep='first')
print("Duplicates have been removed")
print(keepDF)

# Write output to final excel file
print("Start writing to output file...")
filename = "output.xlsx"
writer = pd.ExcelWriter(xlsxFolderPath + "\\" + filename, engine="xlsxwriter")
keepDF.to_excel(writer, sheet_name="main")
mainDF.to_excel(writer, sheet_name="all")
futureDatesDF.to_excel(writer, sheet_name="future_dates")
invalidDatesDF.to_excel(writer, sheet_name="wrong_dates")
writer.save()

# Get all inv._pty
vendor_id_list = keepDF["vendor_id"].unique()

# Write one file per inv._pty
for vendor_id in vendor_id_list:
    filename = "output_{}.xlsx".format(vendor_id)
    writer = pd.ExcelWriter(xlsxFolderPath + "\\" + filename, engine="xlsxwriter")
    keepDF[keepDF['vendor_id'] == vendor_id].to_excel(writer, sheet_name="main")
    writer.save()
    print("Saved output file for vendor_id: {}".format(vendor_id))
